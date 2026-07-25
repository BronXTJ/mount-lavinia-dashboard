"""
Per-GN metrics from Sentinel-2 classified rasters:
built-up %, green %, soft-surface % (+ ha) for y2018 / y2020 / y2025.

Exports plain CSV + styled XLSX (bold, centered, color-coded headers).
"""

from __future__ import annotations

import json
from pathlib import Path

import geopandas as gpd
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import rasterio
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from rasterio.features import geometry_mask
from shapely.geometry import mapping

ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config.json"
PROC = ROOT / "data" / "processed"
TAB = ROOT / "outputs" / "tables"
FIG = ROOT / "outputs" / "figures"
EPOCHS = ["y2018", "y2020", "y2025"]
GN_ORDER = [
    "Mount Lavinia",
    "Kawdana West",
    "Watarappala",
    "Wathumulla",
    "Wedikanda",
]


def load_config() -> dict:
    with CONFIG_PATH.open(encoding="utf-8") as f:
        return json.load(f)


def zone_mask(dataset: rasterio.DatasetReader, geom) -> np.ndarray:
    outside = geometry_mask(
        [mapping(geom)],
        out_shape=(dataset.height, dataset.width),
        transform=dataset.transform,
        invert=False,
        all_touched=True,
    )
    return ~outside


def metrics_for_mask(arr: np.ndarray, mask: np.ndarray, px_m2: float) -> dict:
    valid = mask & (arr >= 1) & (arr <= 5)
    n = int(np.sum(valid))
    if n == 0:
        return {
            "classified_px": 0,
            "area_ha": 0.0,
            "built_up_ha": 0.0,
            "green_ha": 0.0,
            "soft_surface_ha": 0.0,
            "built_up_pct": 0.0,
            "green_pct": 0.0,
            "soft_surface_pct": 0.0,
        }
    built = arr == 1
    green = arr == 2
    soft = (arr == 2) | (arr == 3) | (arr == 4) | (arr == 5)
    built_n = int(np.sum(valid & built))
    green_n = int(np.sum(valid & green))
    soft_n = int(np.sum(valid & soft))
    area_ha = n * px_m2 / 10_000.0
    return {
        "classified_px": n,
        "area_ha": round(area_ha, 4),
        "built_up_ha": round(built_n * px_m2 / 10_000.0, 4),
        "green_ha": round(green_n * px_m2 / 10_000.0, 4),
        "soft_surface_ha": round(soft_n * px_m2 / 10_000.0, 4),
        "built_up_pct": round(100.0 * built_n / n, 2),
        "green_pct": round(100.0 * green_n / n, 2),
        "soft_surface_pct": round(100.0 * soft_n / n, 2),
    }


def compute_table(cfg: dict) -> pd.DataFrame:
    src = (ROOT / cfg["aoi"]["source"]).resolve()
    gdf = gpd.read_file(src)
    if gdf.crs is None:
        gdf = gdf.set_crs(4326)
    gdf = gdf.to_crs("EPSG:32644")

    rows = []
    for epoch in EPOCHS:
        path = PROC / f"classified_{epoch}.tif"
        with rasterio.open(path) as ds:
            arr = ds.read(1)
            px = abs(ds.res[0] * ds.res[1])
            for name in GN_ORDER:
                sub = gdf[gdf["ADM4_EN"] == name]
                if sub.empty:
                    continue
                geom = sub.dissolve().geometry.iloc[0]
                m = zone_mask(ds, geom)
                met = metrics_for_mask(arr, m, px)
                rows.append({"gn": name, "epoch": epoch, **met})
    return pd.DataFrame(rows)


def write_styled_xlsx(df: pd.DataFrame, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "per_gn_metrics"

    fills = {
        "gn": PatternFill("solid", fgColor="D9D9D9"),
        "epoch": PatternFill("solid", fgColor="F2F2F2"),
        "built": PatternFill("solid", fgColor="F4CCCC"),
        "green": PatternFill("solid", fgColor="D9EAD3"),
        "soft": PatternFill("solid", fgColor="CFE2F3"),
        "meta": PatternFill("solid", fgColor="EFEFEF"),
    }
    header_font = Font(bold=True, color="000000")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    def header_fill(col: str) -> PatternFill:
        if col == "gn":
            return fills["gn"]
        if col == "epoch":
            return fills["epoch"]
        if col.startswith("built_up"):
            return fills["built"]
        if col.startswith("green"):
            return fills["green"]
        if col.startswith("soft_surface"):
            return fills["soft"]
        return fills["meta"]

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = center
            cell.border = thin
            if r_idx == 1:
                col_name = df.columns[c_idx - 1]
                cell.font = header_font
                cell.fill = header_fill(col_name)
            else:
                cell.font = Font(bold=False)

    ws.row_dimensions[1].height = 32
    from openpyxl.utils import get_column_letter

    widths = {
        "gn": 16,
        "epoch": 10,
        "classified_px": 14,
        "area_ha": 10,
        "built_up_ha": 12,
        "green_ha": 10,
        "soft_surface_ha": 14,
        "built_up_pct": 12,
        "green_pct": 10,
        "soft_surface_pct": 14,
    }
    for i, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 12)

    # Legend sheet
    legend = wb.create_sheet("legend")
    legend["A1"] = "Column group"
    legend["B1"] = "Meaning"
    legend["A1"].font = header_font
    legend["B1"].font = header_font
    legend["A1"].fill = fills["meta"]
    legend["B1"].fill = fills["meta"]
    rows_leg = [
        ("gn / epoch (gray)", "GN name and epoch id"),
        ("built_up_* (red)", "Class 1 — built-up / impervious"),
        ("green_* (green)", "Class 2 — vegetation / tree cover"),
        ("soft_surface_* (blue)", "Classes 2+3+4+5 — veg + open/bare + water/wetland + beach"),
        ("area_ha / classified_px", "Total classified area inside each GN"),
    ]
    for i, (a, b) in enumerate(rows_leg, start=2):
        legend[f"A{i}"] = a
        legend[f"B{i}"] = b
        legend[f"A{i}"].alignment = center
        legend[f"B{i}"].alignment = Alignment(horizontal="left", vertical="center")
    legend.column_dimensions["A"].width = 28
    legend.column_dimensions["B"].width = 70

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def plot_bars(df: pd.DataFrame, path: Path) -> None:
    fig, axes = plt.subplots(1, 2, figsize=(12, 5), dpi=150)

    d25 = df[df["epoch"] == "y2025"].set_index("gn").loc[GN_ORDER]
    x = np.arange(len(GN_ORDER))
    w = 0.25
    axes[0].bar(x - w, d25["built_up_pct"], width=w, label="Built-up %", color="#d73027")
    axes[0].bar(x, d25["green_pct"], width=w, label="Green %", color="#1a9850")
    axes[0].bar(x + w, d25["soft_surface_pct"], width=w, label="Soft-surface %", color="#4575b4")
    axes[0].set_xticks(x)
    axes[0].set_xticklabels(GN_ORDER, rotation=25, ha="right", fontsize=8)
    axes[0].set_ylabel("% of GN classified area")
    axes[0].set_title("Per-GN metrics — Sentinel-2 ~2025 (10 m)")
    axes[0].legend(fontsize=8)
    axes[0].set_ylim(0, 100)

    d18 = df[df["epoch"] == "y2018"].set_index("gn").loc[GN_ORDER]
    delta_built = d25["built_up_pct"] - d18["built_up_pct"]
    delta_green = d25["green_pct"] - d18["green_pct"]
    delta_soft = d25["soft_surface_pct"] - d18["soft_surface_pct"]
    axes[1].bar(x - w, delta_built, width=w, label="Δ Built-up pp", color="#d73027")
    axes[1].bar(x, delta_green, width=w, label="Δ Green pp", color="#1a9850")
    axes[1].bar(x + w, delta_soft, width=w, label="Δ Soft pp", color="#4575b4")
    axes[1].axhline(0, color="#333", linewidth=0.8)
    axes[1].set_xticks(x)
    axes[1].set_xticklabels(GN_ORDER, rotation=25, ha="right", fontsize=8)
    axes[1].set_ylabel("Percentage-point change")
    axes[1].set_title("Change ~2018 → ~2025 (percentage points)")
    axes[1].legend(fontsize=8)

    fig.tight_layout()
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)


def main() -> None:
    cfg = load_config()
    df = compute_table(cfg)
    # Stable column order
    cols = [
        "gn",
        "epoch",
        "area_ha",
        "classified_px",
        "built_up_ha",
        "built_up_pct",
        "green_ha",
        "green_pct",
        "soft_surface_ha",
        "soft_surface_pct",
    ]
    df = df[cols]

    TAB.mkdir(parents=True, exist_ok=True)
    csv_path = TAB / "per_gn_metrics.csv"
    xlsx_path = TAB / "per_gn_metrics.xlsx"
    fig_path = FIG / "per_gn_built_green_soft_bars.png"

    df.to_csv(csv_path, index=False)
    write_styled_xlsx(df, xlsx_path)
    plot_bars(df, fig_path)

    print(df.to_string(index=False))
    print(f"Wrote {csv_path}")
    print(f"Wrote {xlsx_path}")
    print(f"Wrote {fig_path}")


if __name__ == "__main__":
    main()
